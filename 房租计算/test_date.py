import openpyxl
from datetime import datetime

# 测试Excel文件读取
def test_excel_reading():
    try:
        wb = openpyxl.load_workbook('房间水电气底数.xlsx')
        sheet = wb.active
        print(f"Excel文件读取成功")
        print(f"列数: {sheet.max_column}")
        print(f"行数: {sheet.max_row}")
        
        # 读取第一行数据作为示例
        if sheet.max_row > 1:
            first_row = []
            for cell in sheet[2]:  # 第二行数据
                first_row.append(cell.value)
            print(f"第二行数据: {first_row}")
            
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")

if __name__ == "__main__":
    test_excel_reading() 