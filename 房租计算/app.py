import pandas as pd
from flask import Flask, render_template, request
import openpyxl
from datetime import datetime
import numbers

app = Flask(__name__)

# 自定义过滤器：当小数部分为0时只显示整数
@app.template_filter('smart_float')
def smart_float(value):
    if value is None:
        return '--'
    try:
        float_val = float(value)
        if float_val == int(float_val):
            return str(int(float_val))
        else:
            return f"{float_val:.1f}"
    except:
        return str(value)

# Excel 文件路径
EXCEL_FILE = '房间水电气底数.xlsx'

# 从excel表中读取上个月房租水电气等数据
def get_last_reading(building, room):
    try:        
        wb = openpyxl.load_workbook(EXCEL_FILE) # 使用 openpyxl 读取 Excel 文件
        sheet = wb.active # 获取活动工作表
        building_column = 1  # 楼栋在第1列
        room_column = 2     # 房间号在第2列
        water_column = 3    # 水表数在第3列
        electric_column = 4 # 电表数在第4列
        gas_column = 5      # 气表数在第5列
        water_price_column = 6   # 水单价在第6列
        electric_price_column = 7 # 电单价在第7列
        gas_price_column = 8     # 气单价在第8列
        rent_column = 9          # 房租在第9列
        management_column = 10   # 管理费在第10列
        internet_column = 11     # 网费在第11列
        date_column = 12         # 日期在第12列

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):# 从第二行开始迭代
            # 检查房间号和楼栋是否匹配
            if str(row[room_column - 1]) == str(room) and str(row[building_column - 1]) == str(building):
                return {
                    'building': row[building_column - 1],
                    'water': row[water_column - 1],
                    'electric': row[electric_column - 1],
                    'gas': row[gas_column - 1],
                    'water_price': row[water_price_column - 1],
                    'electric_price': row[electric_price_column - 1],
                    'gas_price': row[gas_price_column - 1],
                    'rent': row[rent_column - 1],
                    'management_fee': row[management_column - 1],
                    'internet_fee': row[internet_column - 1],
                    'date': row[date_column - 1]
                }, sheet, wb
        return None, None, None
    except Exception as e:
        return None, None, None

# 用前端输入的数据更新excel表中的数据
def update_reading(sheet, wb, building, room, current_water, current_electric, current_gas, selected_date=None):
    building_column = 1  # 楼栋列
    room_column = 2  # 房间号列
    water_column = 3  # 水表数列
    electric_column = 4  # 电表数列
    gas_column = 5  # 气表数列
    date_column = 12  # 日期列

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if str(row[room_column - 1].value) == str(room) and str(row[building_column - 1].value) == str(building):
            row[water_column - 1].value = current_water
            row[electric_column - 1].value = current_electric
            row[gas_column - 1].value = current_gas
            # 使用选择的日期，如果没有则使用当前日期
            current_date = selected_date if selected_date else datetime.now().strftime('%Y-%m-%d')
            
            # 更新当前行的所有数据
            sheet.cell(row=row[0].row, column=water_column, value=current_water)
            sheet.cell(row=row[0].row, column=electric_column, value=current_electric)
            sheet.cell(row=row[0].row, column=gas_column, value=current_gas)
            sheet.cell(row=row[0].row, column=date_column, value=current_date)
            wb.save(EXCEL_FILE)
            break

# 计算费用
def calculate_fees(last_water, last_electric, last_gas, current_water, current_electric, current_gas, building, room, last_data):
    # 从Excel数据中获取单价，如果任何一个单价未设置，返回None
    if last_data['water_price'] is None or last_data['electric_price'] is None or last_data['gas_price'] is None:
        return None, None, None, None, None, None, None, None

    water_price = float(last_data['water_price'])
    electric_price = float(last_data['electric_price'])
    gas_price = float(last_data['gas_price'])

    water_usage = current_water - last_water
    electric_usage = current_electric - last_electric
    gas_usage = current_gas - last_gas

    water_fee = round(water_usage * water_price, 2)
    electric_fee = round(electric_usage * electric_price, 2)
    gas_fee = round(gas_usage * gas_price, 2)
    water_usage = int(water_usage)
    electric_usage = int(electric_usage)
    gas_usage = int(gas_usage)

    return water_fee, electric_fee, gas_fee, water_usage, electric_usage, gas_usage, water_price, electric_price

# 计算总费用
def calculate_total(rent, water_fee, electric_fee, gas_fee, management_fee, internet_fee):
    # 确保所有输入都是数值类型
    rent = float(rent) if rent is not None else 0
    water_fee = float(water_fee) if water_fee is not None else 0
    electric_fee = float(electric_fee) if electric_fee is not None else 0
    gas_fee = float(gas_fee) if gas_fee is not None else 0
    management_fee = float(management_fee) if management_fee is not None else 0
    internet_fee = float(internet_fee) if internet_fee is not None else 0
    
    total_price = round(rent + water_fee + electric_fee + gas_fee + management_fee + internet_fee, 2)
    return total_price

def get_all_buildings():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        building_column = 1  # 楼栋在第1列
        buildings = set()
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            val = row[building_column - 1]
            if val is not None and str(val).strip() != '':
                buildings.add(str(val).strip())
        return sorted(buildings)
    except Exception as e:
        return []

@app.route('/')
def index():
    buildings = get_all_buildings()
    # 获取当前日期作为默认日期
    current_date = datetime.now().strftime('%Y-%m-%d')
    return render_template("index.html", buildings=buildings, last_date=current_date)

@app.route('/calculate', methods=['POST'])
def calculate():
    building = request.form['building']
    room = request.form['room']
    current_water = float(request.form['current_water'])
    current_electric = float(request.form['current_electric'])
    current_gas = float(request.form['current_gas'])
    selected_date = request.form.get('selected_date', datetime.now().strftime('%Y-%m-%d'))

    last_data, sheet, wb = get_last_reading(building, room)
    buildings = get_all_buildings()
    if last_data is None:
        return render_template("index.html", total_price=None, room=room, building=building, buildings=buildings, 
                             last_date=selected_date, error="房间号或楼栋不存在或数据读取错误")

    def safe_float(val):
        try:
            if isinstance(val, datetime):
                return 0.0
            return float(val)
        except:
            return 0.0
    last_water = safe_float(last_data['water'])
    last_electric = safe_float(last_data['electric'])
    last_gas = safe_float(last_data['gas'])

    if (current_water < last_water or
        current_electric < last_electric or
        current_gas < last_gas):
        return render_template("index.html", total_price=None, room=room, building=building, buildings=buildings,
                             last_date=selected_date, error="当月水表数、电表数或气表数不能小于上月！")

    result = calculate_fees(
        last_water, last_electric, last_gas,
        current_water, current_electric, current_gas, building, room, last_data)
    
    # 检查是否有未设置的单价
    if result is None or result[0] is None:
        return render_template("index.html", total_price=None, room=room, building=building, buildings=buildings,
                             last_date=selected_date, error="未设置单价，请先在Excel表中设置水电气单价！")
                             
    water_fee, electric_fee, gas_fee, water_usage, electric_usage, gas_usage, water_price, electric_price = result

    total_price = calculate_total(
        last_data['rent'], water_fee, electric_fee, gas_fee,
        last_data['management_fee'], last_data['internet_fee'])

    update_reading(sheet, wb, building, room, current_water, current_electric, current_gas, selected_date)

    return render_template("index.html",
                         total_price=total_price,
                         room=room,
                         building=building,
                         buildings=buildings,
                         water_fee=water_fee,
                         electric_fee=electric_fee,
                         gas_fee=gas_fee,
                         water_usage=water_usage,
                         electric_usage=electric_usage,
                         gas_usage=gas_usage,
                         rent=last_data['rent'],
                         management_fee=last_data['management_fee'],
                         internet_fee=last_data['internet_fee'],
                         last_water=last_water,
                         last_electric=last_electric,
                         last_gas=last_gas,
                         current_water=current_water,
                         current_electric=current_electric,
                         current_gas=current_gas,
                         water_price=water_price,
                         electric_price=electric_price,
                         date=selected_date,
                         last_date=selected_date)

if __name__ == '__main__':
    app.run(debug=True)